#! /usr/bin/env python
# import binascii
# -*-coding:UTF-8-*-
import os
import sys
import openpyxl

sys.path.append(os.path.join(os.path.dirname(__file__), ".."))

class DataProcess:

    def __init__(self):
        print("[DataFormatConversion]:initial")

        self.DataFormatConversion(True)

    def DataFormatConversion(self,PrintEnable):
        wb=openpyxl.load_workbook('PortAndSignal.xlsx')
        # wb=openpyxl.Workbook()
        sheet=wb.active
        if 'ProcessData1' not in wb.sheetnames:wb.create_sheet('ProcessData1')
        sheet['A1']='ID_xx_E'
        sheet['B1']='KeyName'
        sheet['C1']='Note'

        data=[('Alice',25),('Bob',30),('Carol',28)]
        for row_index,(name,age) in enumerate(data,start=2):
            sheet[f'A{row_index}']=name
            sheet[f'B{row_index}']=age

        with open('PortAndSignal.txt', 'r') as file:
            for line in file:
                # print(line.strip())
                lines=line.strip('--!')
                # print(lines.split())
                lines_split=lines.split()
                lines_split_len=len(lines_split)
                print(lines_split_len)
                print(lines_split[0])
                # for i in range(lines_split_len):
                #     sheet['A[%s]']=lines_split[i] %(i+2)
        
        if 'ProcessData2' not in wb.sheetnames:wb.create_sheet('ProcessData2')
        if 'ProcessData3' not in wb.sheetnames:wb.create_sheet('ProcessData3')
        wb.save('PortAndSignal.xlsx')

        if PrintEnable==True:
            print("Please define the infomation of printing that you need!!!")

            
if __name__ == '__main__':
    gen=DataProcess()