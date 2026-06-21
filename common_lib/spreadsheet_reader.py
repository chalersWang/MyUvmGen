#!/usr/bin/env python3
# -*-coding:UTF-8-*-
# ============================================================
# MyUvmGen_v2.0_macos — 统一电子表格读取层
# 支持 .xlsx (openpyxl) 和 .numbers (AppleScript+Numbers.app)
# ============================================================
import os, sys, tempfile, subprocess, shutil

class NumbersCell:
    __slots__ = ('value', 'row', 'column')
    def __init__(self, value, row, column):
        self.value = value; self.row = row; self.column = column

class NumbersSheet:
    """模拟 openpyxl Worksheet"""
    def __init__(self, title, rows_data):
        self.title = title
        self._rows = [[v for v in r] for r in rows_data]
        self.max_column = max((len(r) for r in rows_data), default=0)
        self.max_row = len(rows_data)
        self._cols = {}
        for ri, row in enumerate(rows_data, 1):
            for ci, val in enumerate(row, 1):
                letter = _c2l(ci)
                self._cols.setdefault(letter, []).append(NumbersCell(val, ri, ci))
    def __getitem__(self, col_letter):
        return self._cols.get(col_letter, [])
    def iter_rows(self, min_row=1, max_row=None, min_col=1, max_col=None, values_only=True):
        mr = min(max_row or self.max_row, self.max_row)
        mc = min(max_col or self.max_column, self.max_column)
        for ri in range(min_row-1, mr):
            row = self._rows[ri] if ri < len(self._rows) else []
            if values_only:
                yield tuple(row[ci] if ci < len(row) else None for ci in range(min_col-1, mc))
            else:
                yield tuple(NumbersCell(row[ci] if ci < len(row) else None, ri+1, ci+1) for ci in range(min_col-1, mc))

class NumbersWorkbook:
    def __init__(self):
        self._sheets = {}; self.sheetnames = []
    def __getitem__(self, name):
        return self._sheets[name]

def _c2l(idx):
    s = ''
    while idx > 0: idx, r = divmod(idx-1, 26); s = chr(65+r) + s
    return s

def _auto(v):
    """智能类型转换"""
    if v is None: return None
    try:
        if '.' in str(v): return float(v)
        return int(v)
    except: return v

def load_workbook(filepath, data_only=False):
    ext = os.path.splitext(filepath)[1].lower()
    if ext in ('.xlsx', '.xls'):
        import openpyxl
        return openpyxl.load_workbook(filepath, data_only=data_only)
    elif ext == '.numbers':
        return _load_numbers(filepath)
    else:
        raise ValueError("不支持格式: %s" % ext)

def _load_numbers(filepath):
    if sys.platform != 'darwin':
        raise OSError("读取 .numbers 需要 macOS")
    
    abs_path = os.path.abspath(filepath)
    if not os.path.exists(abs_path):
        raise FileNotFoundError("文件不存在: %s" % abs_path)
    
    # AppleScript 读取 Numbers 数据
    # 用 ASCII 分隔符 SOH (0x01) 作为单元格分隔
    SEP = chr(1)
    
    ascript = '''on run {inputPath}
    tell application "Numbers"
        set theDoc to open file (POSIX file inputPath as text)
        set output to "SHEETS:" & return
        
        repeat with aSheet in sheets of theDoc
            set output to output & "SHEET:" & name of aSheet & return
            repeat with aTable in tables of aSheet
                set output to output & "TABLE:" & name of aTable & return
                set rowIdx to 0
                repeat with aRow in rows of aTable
                    set rowIdx to rowIdx + 1
                    if rowIdx > 500 then exit repeat
                    set cellLine to ""
                    repeat with aCell in cells of aRow
                        set cv to value of aCell
                        if cv is missing value then
                            set cellLine to cellLine & "''' + SEP + '''"
                        else
                            try
                                set cvStr to cv as text
                                set cellLine to cellLine & cvStr & "''' + SEP + '''"
                            on error
                                set cellLine to cellLine & "''' + SEP + '''"
                            end try
                        end if
                    end repeat
                    set output to output & cellLine & return
                end repeat
                exit repeat
            end repeat
        end repeat
        
        close theDoc without saving
        return output
    end tell
    end run'''
    
    proc = subprocess.run(['osascript', '-e', ascript, abs_path],
                          capture_output=True, text=True, timeout=120)
    
    if proc.returncode != 0:
        raise RuntimeError("AppleScript错误:\n" + proc.stderr)
    
    # 解析
    sheets = []
    current_sheet, current_rows = None, []
    
    for line in proc.stdout.split('\n'):
        line = line.strip()
        if not line or line == 'SHEETS:':
            continue
        if line.startswith('SHEET:'):
            if current_sheet is not None and current_rows:
                sheets.append((current_sheet, current_rows))
            current_sheet = line[6:].strip()
            current_rows = []
        elif line.startswith('TABLE:'):
            continue
        else:
            cells = [_auto(c) if c else None for c in line.split(SEP)]
            while cells and cells[-1] is None:
                cells.pop()
            if cells:
                current_rows.append(cells)
    
    if current_sheet is not None and current_rows:
        sheets.append((current_sheet, current_rows))
    
    if not sheets:
        raise RuntimeError("未读取到数据，请确认 .numbers 文件有表格内容")
    
    wb = NumbersWorkbook()
    for sname, rows in sheets:
        ns = NumbersSheet(sname, rows)
        wb._sheets[sname] = ns
        wb.sheetnames.append(sname)
    
    return wb


if __name__ == '__main__':
    wb = load_workbook(sys.argv[1])
    print("Sheets:", wb.sheetnames)
    for sn in wb.sheetnames[:5]:
        s = wb[sn]
        print("\n--- %s (%d rows) ---" % (sn, s.max_row))
        for row in s.iter_rows(max_row=min(s.max_row,10)):
            print('  | '.join(str(v)[:60] if v is not None else '' for v in row))
