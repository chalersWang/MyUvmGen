# -*-coding:UTF-8-*-
"""自动转换 .numbers → .xlsx"""
import os, sys, subprocess

def convert_numbers_if_needed(filepath):
    if not filepath.endswith('.numbers'):
        return filepath
    
    xlsx_path = filepath.replace('.numbers', '.xlsx')
    need_convert = False
    
    if not os.path.exists(xlsx_path):
        need_convert = True
        print("[Auto Convert] .xlsx 不存在，需要转换 .numbers")
    else:
        numbers_mtime = os.path.getmtime(filepath)
        xlsx_mtime = os.path.getmtime(xlsx_path)
        if numbers_mtime > xlsx_mtime:
            need_convert = True
            print("[Auto Convert] .numbers 已更新，重新转换 .xlsx")
        else:
            print("[Auto Convert] .xlsx 已是最新，跳过转换")
            return xlsx_path
    
    if not need_convert:
        return xlsx_path
    
    print("[Auto Convert] 正在转换 %s → %s ..." % (filepath, xlsx_path))
    
    # 方式1: numbers-parser
    try:
        from numbers_parser import Document
        import openpyxl as px
        doc = Document(filepath)
        wb = px.Workbook()
        wb.remove(wb.active)
        for s in doc.sheets:
            ws = wb.create_sheet(title=s.name)
            for table in s.tables:
                for ri, row in enumerate(table.rows(), 1):
                    for ci, cell in enumerate(row, 1):
                        ws.cell(row=ri, column=ci, value=cell.value)
        wb.save(xlsx_path)
        print("[Auto Convert] OK via numbers-parser")
        return xlsx_path
    except (ImportError, Exception) as e:
        print("[Auto Convert] numbers-parser skip: %s" % str(e)[:80])
    
    # 方式2: AppleScript
    if sys.platform == 'darwin':
        script = 'tell application "Numbers" to open POSIX file "%s"\n' % os.path.abspath(filepath)
        script += 'tell application "Numbers" to export document 1 to POSIX file "%s" as Microsoft Excel\n' % os.path.abspath(xlsx_path)
        script += 'tell application "Numbers" to close document 1 without saving\n'
        
        try:
            proc = subprocess.run(['osascript', '-e', script], capture_output=True, text=True, timeout=300)
            if os.path.exists(xlsx_path) and os.path.getsize(xlsx_path) > 0:
                print("[Auto Convert] OK via Numbers.app")
                return xlsx_path
            else:
                print("[Auto Convert] Numbers.app failed: %s" % proc.stderr[:200])
        except Exception as e:
            print("[Auto Convert] AppleScript error: %s" % str(e)[:200])
    
    print("[Auto Convert] FAILED. Please run: ./convert_numbers.sh %s" % filepath)
    raise RuntimeError("Cannot convert .numbers automatically. Use convert_numbers.sh")

